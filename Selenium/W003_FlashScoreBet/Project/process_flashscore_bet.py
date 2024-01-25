'''
The process itself and logic for process
'''

import os
import logging
import datetime

import openpyxl
import pandas as pd

# library of my custom methods
from Libraries import utilities
from Libraries.framework import FrameWork
from Libraries.flashscore import FlashScore


# Check if deactivated league are deactivated more than absence days
def prepare_input_file(config_dict: dict) -> None:
    try:

        logging.info(f"Method prepare_input_file started")
        if len(config_dict) > 0:
            # Create instance of FrameWork
            framework_instance = FrameWork()
            config_dict = framework_instance.config_dict

        input_file_path = config_dict["input_file_path"]
        absence_days = config_dict["absence_days"]

        # load excel
        df_excel = pd.read_excel(input_file_path)

        # Calc number of days between date column and today
        df_excel["days"] = (datetime.datetime.today() - df_excel["date"]).dt.days

        # If number days between today and last check is less the abcence days retrive item to be checked
        df_excel.active[df_excel["days"] > int(absence_days)] = True
        df_excel.date[df_excel["days"] > int(absence_days)] = ""

        # Delete column days
        df_excel.drop(["days"], axis=1, inplace=True)
        # Save data in excel
        df_excel.to_excel(input_file_path, index=False)
    except Exception as error_message:
        raise Exception(f"Method 'prepare_input_file'  fails - {error_message}") from error_message


# Delete previous file
def prepare_output_file(config_dict: dict) -> None:
    try:
        logging.info(f"Method prepare_output_file started")
        if len(config_dict) > 0:
            # Create instance of FrameWork
            framework_instance = FrameWork()
            config_dict = framework_instance.config_dict

        # Get output file path
        output_file_path = config_dict["output_file_path"]

        # Delete current output file
        if os.path.exists(output_file_path):
            os.remove(output_file_path)

    except Exception as error_message:
        raise Exception(f"Method 'prepare_output_file'  fails - {error_message}") from error_message


# Convert Excel file into prety view
def adjust_excel_columns(file_path: str) -> None:
    # Open excel
    excel_output = openpyxl.load_workbook(file_path)
    sheet = excel_output.active

    # Adjust column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width
    # save excel
    excel_output.save(file_path)


# Filter data based on provided criteria
def reorder_output_table(file_path: str) -> None:
    try:
        # Read Excel to Dataframe
        df_table = pd.read_excel(file_path)
        # Reorder columns
        df_table = df_table[
            ["type", "position", "club_name", "percentage_of_wins", "percentage_of_loses", "next_game","country", "league",
             "matches_played", "points", "wins", "draws", "loses", "balance", "goals_difference", "matches_all",
             "points_left", "points_to_first_place", "points_differece", "points_to_champion"]]

        # Save DataFrame in Excel
        df_table.to_excel(file_path, index=False)

    except Exception as error_message:
        raise Exception(f"Method 'reorder_output_table' fails - {error_message}") from error_message


# Adjust column for Input and Output files
def adjust_final_excels(config_dict: dict) -> None:
    try:
        if len(config_dict) > 0:
            # Create instance of FrameWork
            framework_instance = FrameWork()
            config_dict = framework_instance.config_dict

        # Get input and output files path
        input_file_path = config_dict["input_file_path"]
        output_file_path = config_dict["output_file_path"]

        # Reorder columns
        reorder_output_table(output_file_path)

        # Adjust columns
        adjust_excel_columns(input_file_path)
        adjust_excel_columns(output_file_path)

    except Exception as error_message:
        raise Exception(f"Method 'adjust_final_excels'  fails - {error_message}") from error_message


# Get data from flashscore
def get_table_from_flashscore(in_url: str, in_webdriver_path) -> pd.DataFrame:
    try:
        # Create Instance
        flashscore_instance = FlashScore()

        # Open url
        flashscore_instance.open_url(in_url, in_webdriver_path)

        # Get data from table
        df_table = flashscore_instance.get_main_table_data()

        # Get country & league
        country_name = flashscore_instance.get_country_name()
        league_name = flashscore_instance.get_league_name()
        ###Move country & league to add additional data???

        # Add additional calculations
        df_table = add_additional_data_to_table(df_table, country_name, league_name)

        return df_table

    except Exception as error_message:
        raise Exception(f"Method 'get_table_from_flashscore' fails - {error_message}") from error_message


# Add additional columns with extra calculations
def add_additional_data_to_table(df_table: object, country_name: str = "espana",
                                 league_name: str = "la liga") -> object:
    # DEBUG
    # Read dataframe from file
    # df_table = pd.read_excel("Test/TEST.xlsx")

    try:
        # Save country and league
        df_table["country"] = country_name
        df_table["league"] = league_name

        # Percentage of wins & loses
        df_table["percentage_of_wins"] = round(df_table["wins"] / df_table["matches_played"], 2)
        df_table["percentage_of_loses"] = round(df_table["loses"] / df_table["matches_played"], 2)

        # Matched all & matched left
        df_table["matches_all"] = (df_table.shape[0] - 1) * 2
        df_table["points_left"] = (df_table["matches_all"] - df_table["matches_played"]) * 3
        # Points to first place
        df_table["points_to_first_place"] = df_table.loc[0:, "points"] - df_table.at[0, "points"]
        # Points difference
        df_table["points_differece"] = df_table["points"].diff()
        # Points to champion
        df_table["points_to_champion"] = 0
        df_table.at[0, "points_to_champion"] = df_table.at[0, "points_left"] - df_table.at[1, "points_differece"]

        return df_table
    except Exception as error_message:
        raise Exception(f"Method 'add_additional_data_to_table' fails - {error_message}") from error_message


# Filter data based on provided criteria
def filter_table(df_table: object, percentage_cut: float = 0.8,
                 matches_played_cut: int = 10) -> object:
    try:
        # Filter teams that played more than 10 games
        df_table = df_table[df_table["matches_played"] >= matches_played_cut]

        # Get clubs that meet criteria for win & loss percentage (0.80)
        df_wins = df_table[df_table["percentage_of_wins"] >= percentage_cut]
        df_wins["type"] = "winner"
        df_loses = df_table[df_table["percentage_of_loses"] >= percentage_cut]
        df_loses["type"] = "loser"

        # Merge data
        df_all = pd.concat([df_wins, df_loses], axis=0)

        return df_all

    except Exception as error_message:
        raise Exception(f"Method 'filter_table' fails - {error_message}") from error_message


# Process all rows in input file
def process_input_file(config_dict: dict = None) -> None:
    try:
        if config_dict == None:
            # Create instance of FrameWork
            framework_instance = FrameWork()
            config_dict = framework_instance.config_dict

        input_file_path = config_dict["input_file_path"]
        output_file_path = config_dict["output_file_path"]
        webdriver_path = config_dict["webdriver_path"]
        percentage_cut = float(config_dict["percentage_cut"])
        matches_played_cut = int(config_dict["matches_played_cut"])

        # load excel
        excel_input = openpyxl.load_workbook(input_file_path)
        sheet = excel_input.active

        # iterate through excel and display data
        for i in range(2, sheet.max_row + 1):
            # print progress
            print(f"Progress: {str(i)}/{str(sheet.max_row)} - {'{0:.2f}'.format(i / sheet.max_row * 100)}%")

            # Get data from excel
            active_cell_value = bool(sheet.cell(row=i, column=3).value)
            url = sheet.cell(row=i, column=1).value

            # if not active get another
            if not active_cell_value:
                continue

            # Get table from flashscore
            try:
                df_flashscore = get_table_from_flashscore(url, webdriver_path)
            except Exception as error_message:
                sheet.cell(row=i, column=2).value = "???"
                sheet.cell(row=i, column=7).value = str(error_message)
                continue

            # Sort values by percentage acending
            df_flashscore.sort_values(by=['percentage_of_wins'], ascending=False)
            # Save high_score
            high_score = df_flashscore.at[0, 'percentage_of_wins']
            sheet.cell(row=i, column=2).value = high_score

            # If high_score less than 0.65 - deactivate
            if high_score < (percentage_cut - 0.15):
                # Set active cell as False
                sheet.cell(row=i, column=3).value = False
                sheet.cell(row=i, column=5).value = datetime.datetime.today()

            # Filter tale based on provided criteria | Matches played >10; Wins/losses% >0.8
            df_flashscore = filter_table(df_flashscore, percentage_cut, matches_played_cut)

            # Append flashscore data to output excel
            utilities.append_to_excel(output_file_path, df_flashscore)

            # Save progress in excel
            excel_input.save(input_file_path)
    except Exception as error_message:
        raise Exception(f"Method 'process_input_file'  fails - {error_message}") from error_message


# The logic of whole process
def process() -> None:
    try:
        # Create instance of FrameWork
        framework_instance = FrameWork()
        # Set config
        config_dict = framework_instance.config_dict

        # Prepare Input File - Check deactivated rows
        prepare_input_file(config_dict)
        # Prepare output file - Delete previous one
        prepare_output_file(config_dict)

        # Process all items in input file
        process_input_file(config_dict)

        # Adjust column width in files
        adjust_final_excels(config_dict)

        # Get process duration
        process_duration = framework_instance.get_process_time()
        print(f"Process finished. Time: {process_duration}")
        logging.info(f"Time - {process_duration}")

    except Exception as error_message:
        logging.info(f"Main: Unexpected error- '{error_message}'")

# # DEBUG - Local testing
