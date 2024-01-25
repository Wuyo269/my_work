import os
import time
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.chrome.service import Service
from webdrivermanager.chrome import ChromeDriverManager
import pandas as pd
import openpyxl
import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service


'''INPUTS'''
input_file_name = "Input_Flashscore.xlsx"
output_file_name = "Output_Flashscore.xlsx"
webdriver_path = "C:\Program Files (x86)\chromedriver.exe"
url_elecion = r"https://wybory.gov.pl/sejmsenat2023/pl/sejm/wynik/pl"
url = url_elecion



#DONE
def unify_names(in_df,in_column_name):
    #Create a dictionary of unified names
    dict_unified_names = {}
    dict_unified_names['KOMITET WYBORCZY PRAWO I SPRAWIEDLIWOŚĆ'] = 'PIS'
    dict_unified_names['KOALICYJNY KOMITET WYBORCZY KOALICJA OBYWATELSKA PO .N IPL ZIELONI'] = 'PO'
    dict_unified_names['KOALICYJNY KOMITET WYBORCZY TRZECIA DROGA POLSKA 2050 SZYMONA HOŁOWNI - POLSKIE STRONNICTWO LUDOWE'] = 'HOŁOWNIA'
    dict_unified_names['KOMITET WYBORCZY NOWA LEWICA'] = 'LEWICA'
    dict_unified_names['KOMITET WYBORCZY KONFEDERACJA WOLNOŚĆ I NIEPODLEGŁOŚĆ'] = 'KONFEDERACJA'
    dict_unified_names['KOMITET WYBORCZY BEZPARTYJNI SAMORZĄDOWCY'] = 'BEZPARTYJNI'
    dict_unified_names['KOMITET WYBORCZY POLSKA JEST JEDNA'] = 'PJJ'
    dict_unified_names['KW BEZPARTYJNI SAMORZĄDOWCY'] = 'BEZPARTYJNI'
    dict_unified_names['KKW TRZECIA DROGA PSL-PL2050 SZYMONA HOŁOWNI'] = 'HOŁOWNIA'
    dict_unified_names['KW NOWA LEWICA'] = 'LEWICA'
    dict_unified_names['KW PRAWO I SPRAWIEDLIWOŚĆ'] = 'PIS'
    dict_unified_names['KW KONFEDERACJA WOLNOŚĆ I NIEPODLEGŁOŚĆ'] = 'KONFEDERACJA'
    dict_unified_names['KOALICYJNY KOMITET WYBORCZY KOALICJA OBYWATELSKA PO .N IPL ZIELONI'] = 'PO'
    dict_unified_names['KKW KOALICJA OBYWATELSKA PO .N IPL ZIELONI'] = 'PO'
    dict_unified_names['KW POLSKA JEST JEDNA'] = 'PJJ'

    #Check if Item in dict
    for item in in_df[in_column_name].unique():
        if not item in dict_unified_names:
            print(f"dict_unified_names['{item}'] = 'ADDNAME'")

    #unify names
    for item in dict_unified_names:
        in_df[in_column_name] = in_df[in_column_name].replace([item], dict_unified_names[item])

    return in_df
#DONE
#Open elecion url | return driver
def open_eleccion_url(in_url,in_webdriver_path):
    #DEBUG
    #in_url = 'https://wybory.gov.pl/sejmsenat2023/pl/sejm/wynik/okr/2'

    # Create webdriver object
    service = Service(in_webdriver_path)
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)

    #open URL
    driver.get(in_url)

    return driver

# Get list of persons who got mandate | the winners of district
#DONE
def get_district_winners(in_driver):
    # Get element ul
    element_ul = WebDriverWait(in_driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#root > div.res > div.can.mt-3 > ul")))

    # Get element li
    li_elements = element_ul.find_elements(By.TAG_NAME, "li")

    # Initialize empty array for data
    data = []

    #loop through all li elements
    for li in li_elements:
        #Get party name
        party_name = li.find_element(By.TAG_NAME, 'a').text
        # Get sub element li
        li_sub_elements = li.find_elements(By.TAG_NAME, "li")
        # loop through all li elements
        for li in li_sub_elements:
            #Get person name
            person_name = li.find_element(By.TAG_NAME, 'a').text
            #Add data t array
            data.append([person_name,party_name])

    #Create panda dataframe
    df_district_winners = pd.DataFrame(data,columns=["person_name","party_name"])
    #Unify names
    df_district_winners = unify_names(in_df=df_district_winners, in_column_name='party_name')

    return df_district_winners

#DONE
#Get parties summary | How many votes each party got
def get_parties_summary(in_driver):
    # Get summary table
    summary_table_element = WebDriverWait(in_driver, 20).until(
        EC.presence_of_element_located((By.ID, "DataTables_Table_0")))

    #Get element tbody
    element_tbody = WebDriverWait(summary_table_element, 20).until(
        EC.presence_of_element_located((By.TAG_NAME, "tbody")))

    # Get all element tr
    elements_tr = element_tbody.find_elements(By.TAG_NAME, "tr")

    #Initialize empty array for data
    data = []

    #Loop throgh each row to get all summary data
    for row_tr in elements_tr:
        row_elements = row_tr.find_elements(By.TAG_NAME, "td")
        party_name = row_elements[0].text
        votes_numer = int(row_elements[1].text.replace(" ",""))
        votes_percentage = float(row_elements[2].text.replace(",",".").strip('%'))/100
        mandates_numer = int(row_elements[3].text.replace(" ", ""))
        mandates_percentage = float(row_elements[4].text.replace(",",".").strip('%'))/100
        #Add data to array
        data.append([party_name, votes_numer,votes_percentage,mandates_numer,mandates_percentage])

    #Create panda dataframe
    df_parties_summary = pd.DataFrame(data,columns=["party_name", "votes_number","votes_percentage", "mandates_number",
                                                    "mandates_percentage"])
    #Unify names
    df_parties_summary = unify_names(in_df=df_parties_summary, in_column_name='party_name')

    return df_parties_summary

#DONE
#Get party details | Get list of all party members and theirs votes
def get_party_details(in_driver):


    #Get web element table-responsive
    committee_details_tables = in_driver.find_elements(By.CLASS_NAME, 'col-xs-12.col-xl-6.table-responsive')

    # Initialize empty array for data
    data = []

    #loop through all committees
    for committee_element in committee_details_tables:
        #get element h5
        h5_element = committee_element.find_element(By.TAG_NAME, "h5")

        #get the class name of element h5
        h5_class_name = h5_element.get_attribute("class")
        #if class name is strikeout committee is disabled | get next one
        if h5_class_name == "strikeout":
            continue

        #Get committee name
        committee_name = h5_element.find_element(By.TAG_NAME, "a").text

        # Get district name
        element_header = WebDriverWait(in_driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='res'] div[class='row'] h3")))
        district_name = str(element_header.text).split(" ")[-1]

        #Get tbody element
        element_tbody =WebDriverWait(committee_element, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, "tbody")))
        #Get tr elements. These are rows of persons
        elements_tr = element_tbody.find_elements(By.TAG_NAME, 'tr')

        #loop through all persons on list
        for tr_item in elements_tr:
            # if class name is strikeout person is disabled | go to next one
            if tr_item.get_attribute("class") == "strikeout":
                continue

            #Get td elements = columns
            elements_td = tr_item.find_elements(By.TAG_NAME, "td")
            #Get details
            position_on_list = int(elements_td[0].text)
            person_name = elements_td[1].text
            votes_number = int(elements_td[2].text.replace(" ",""))
            #Add to array
            data.append([person_name, position_on_list, votes_number,committee_name,"Distric " +str(district_name)])

    # Create panda dataframe
    df_party_details = pd.DataFrame(data,columns=["person_name", "position_on_list", "votes_number", "party_name", "district_name"])
    #Unify names
    df_party_details = unify_names(in_df=df_party_details, in_column_name='party_name')

    return df_party_details

#DOne
#Get district mandates number | how many persons are elected from this district
def get_district_mandates_number(in_driver):
    # Get table of the summary of all parties
    parties_summary = WebDriverWait(in_driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table[id='DataTables_Table_0'] tfoot tr")))

    #The number of mandates
    district_mandates_number = parties_summary.text.split("\n")[2]

    #Create panda dataframe
    df_district_mandates_number = pd.DataFrame([district_mandates_number],columns=["district_mandates_number"])

    return df_district_mandates_number

#DONE
#Get links to each district details information
def get_electoral_districts_links(in_webdriver_path, in_url):
    #Create empty dictionary
    dict_electoral_districts = {}

    # Create webdriver object
    # driver = webdriver.Chrome(webdriver_path)
    service = Service(in_webdriver_path)
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)

    #open URL
    driver.get(in_url)

    # Get table all Okregi wyborcze
    table_electoral_districts = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CLASS_NAME, "columns3")))

    # Get Rows - All electorals districts
    rows_li = table_electoral_districts.find_elements(By.TAG_NAME, "li")

    #For every electorl district in rows_li
    for row in rows_li:
        electoral_district_path = row.find_element(By.TAG_NAME, "a").get_attribute("href")
        electoral_district_name = row.find_element(By.TAG_NAME, "a").text
        dict_electoral_districts[electoral_district_name] = electoral_district_path

    return  dict_electoral_districts

#Get details information from disrict
def get_electoral_districts_details(in_webdriver_path, in_url):
    #Create empty dictionary
    dict_electoral_district_information = {}
    #DEBUG
    #in_url = 'https://wybory.gov.pl/sejmsenat2023/pl/sejm/wynik/okr/2'

    #open URL
    driver = open_eleccion_url(in_url=in_url, in_webdriver_path=in_webdriver_path)

    #Get DataFrame Data
    df_district_winners = get_district_winners(in_driver=driver)
    df_district_mandates_number = get_district_mandates_number(in_driver=driver)
    df_parties_summary = get_parties_summary(in_driver=driver)
    df_party_details = get_party_details(in_driver=driver)

    #Assign data to dictionary
    dict_electoral_district_information["df_district_winners"] = df_district_winners
    dict_electoral_district_information["df_total_number_of_mandates"] = df_district_mandates_number
    dict_electoral_district_information["df_parties_summary"] = df_parties_summary
    dict_electoral_district_information["df_party_details"] = df_party_details

    return dict_electoral_district_information

#Get details information from disrict
def get_electoral_districts_details2(in_webdriver_path, in_url):
    #Create empty dictionary
    dict_electoral_district_information = {}
    #DEBUG
    #in_url = 'https://wybory.gov.pl/sejmsenat2023/pl/sejm/wynik/okr/2'

    # Create webdriver object
    service = Service(in_webdriver_path)
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)

    #open URL
    driver.get(in_url)

    #Total number of mandates
    ########################################################################################################################
    # Get table all Okregi wyborcze
    committee_summary = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table[id='DataTables_Table_0'] tfoot tr")))

    #The number of mandates
    total_numer_of_mandates = committee_summary.text.split("\n")[2]

    #Create panda dataframe
    data = [total_numer_of_mandates]
    df_total_numer_of_mandates = pd.DataFrame(data,columns=["total_number_of_mandates"])

    dict_electoral_district_information["df_total_number_of_mandates"] = df_total_numer_of_mandates
    ########################################################################################################################

    #Party Summary
    ########################################################################################################################

    # Get table all Okregi wyborcze
    committee_summary = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "DataTables_Table_0")))

    committee_summary_tbody = WebDriverWait(committee_summary, 20).until(
        EC.presence_of_element_located((By.TAG_NAME, "tbody")))

    # Get table all Okregi wyborcze
    committee_summary_tr = committee_summary_tbody.find_elements(By.TAG_NAME, "tr")

    #Initialize empty array for data
    data = []

    #Loop throgh each row to get all summary data
    for row_tr in committee_summary_tr:
        row_elements = row_tr.find_elements(By.TAG_NAME, "td")
        party_name = row_elements[0].text
        votes_numer = int(row_elements[1].text.replace(" ",""))
        votes_percentage = float(row_elements[2].text.replace(",",".").strip('%'))/100
        mandates_numer = int(row_elements[3].text.replace(" ", ""))
        mandates_percentage = float(row_elements[4].text.replace(",",".").strip('%'))/100
        data.append([party_name, votes_numer,votes_percentage,mandates_numer,mandates_percentage])

    #Create panda dataframe
    df_data_parties_summary = pd.DataFrame(data,columns=["party_name", "votes_number","votes_percentage", "mandates_number","mandates_percentage"])
    #Unify names
    df_data_parties_summary = unify_names(in_df=df_data_parties_summary, in_column_name='party_name')

    #DEBUG
    #for item in data:
    #    print(item[0], " - ", item[1], " - ", '{:.2%}'.format(item[1]/sum(df_data_summary["Votes"])))
    #print(sum(df_data_summary["Votes"]))


    #Add data to dictionary
    dict_electoral_district_information["df_data_parties_summary"] = df_data_parties_summary
    ########################################################################################################################

    #Party Details
    ########################################################################################################################

    #Get web element committee tables
    committee_details_tables = driver.find_elements(By.CLASS_NAME, 'col-xs-12.col-xl-6.table-responsive')

    # Initialize empty array for data
    data = []

    #loop through all committees
    for committee_element in committee_details_tables:

        #get element h5
        h5_element = committee_element.find_element(By.TAG_NAME, "h5")

        #get the class name of element h5
        h5_class_name = h5_element.get_attribute("class")
        #if class name is strikeout committee is disabled
        if h5_class_name == "strikeout":
            continue

        #Get committee name
        party_name = h5_element.find_element(By.TAG_NAME, "a").text
        print(party_name)

        #Get tbody element
        tbody_element =WebDriverWait(committee_element, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, "tbody")))

        #Get tr elements. These are rows of persons
        tr_elements = committee_details_tables = tbody_element.find_elements(By.TAG_NAME, 'tr')

        #loop through all persons on list
        for tr_item in tr_elements:

            # if class name is strikeout person is disabled
            if tr_item.get_attribute("class") == "strikeout":
                continue

            #Get td elements = columns
            td_elements = tr_item.find_elements(By.TAG_NAME, "td")
            #Get details
            position_on_list = int(td_elements[0].text)
            person_name = td_elements[1].text
            person_votes_number = int(td_elements[2].text.replace(" ",""))
            #Add to array
            data.append([person_name, position_on_list, person_votes_number,party_name,"Distric " +in_url.split("/")[-1]])


    # Create panda dataframe
    df_data_party_details = pd.DataFrame(data,columns=["person_name", "position_on_list", "person_votes_number", "party_name", "district_name"])
    #Unify names
    df_data_party_details = unify_names(in_df=df_data_party_details, in_column_name='party_name')

    #DEBUG
    #print(df_data_party_summary[df_data_party_summary.person_votes_number == df_data_party_summary.person_votes_number.max()])
    #print(df_data_party_summary.person_votes_number.max())

    #Add data to dictionary
    dict_electoral_district_information["df_data_party_details"] = df_data_party_details
    ########################################################################################################################

    return dict_electoral_district_information

#Save all DataFrames in dictionary to Excel File
def save_df_to_excel(excel_name,dict_of_dataframes):
    #If excel name does not have extenstion - add it
    if not str(excel_name).endswith(".xlsx"):
        excel_name += ".xlsx"
    # 1. Create a pandas excel writer instance and name the Excel file
    xlwriter = pd.ExcelWriter(excel_name)
    #NB: If you don't include a file path like 'C:\Users\Ron\Desktop\File_Name.xlsx'
    # It will save to your default folder, that is,
    #where the file you're reading from is located.

    # 2. Write each dataframe to a worksheet with a name
    for item in dict_of_dataframes:
        df_item = dict_of_dataframes[item]
        df_item.to_excel(xlwriter, sheet_name=item, index=False)

    # 3. Close the instance
    xlwriter.close()

#Load data saved in Excel file to dict of DataFrames
def load_df_from_excel(in_excel_path):
    #Initialize dictionary
    dict_electoral_district_information = {}

    #Create excel object
    xls = pd.ExcelFile(in_excel_path)
    #Loop through all sheets and assign them to dictionary
    for sheet_name in xls.sheet_names:
        dict_electoral_district_information[sheet_name] = pd.read_excel(xls, sheet_name)

    return  dict_electoral_district_information


url_test = 'https://wybory.gov.pl/sejmsenat2023/pl/sejm/wynik/okr/2'

# Get details from current district
dict_electoral_district_information = get_electoral_districts_details(in_webdriver_path=webdriver_path,
                                                                      in_url=url_test)

excel_path = r"/tensorEnv/Selenium/Los Eleciones/Input/Okręg wyborczy nr 1.xlsx"

dict_electoral_district_information = load_df_from_excel(in_excel_path=excel_path)


#System D'Hondt
df_total_number_of_mandates = dict_electoral_district_information["total_number_of_mandates"]
in_mandates_in_district = df_total_number_of_mandates['total_numer_of_mandates'].iloc[0]

in_df_parties_summary = dict_electoral_district_information["df_data_parties_summary"]
in_df_data_party_details = dict_electoral_district_information["df_data_party_details"]






#Create empty array
data = []

#Loop through all parties
for y in range(in_df_parties_summary.shape[0]):
    #Get Party name and votes
    party_name = in_df_parties_summary["party_name"].iloc[y]
    votes_number = in_df_parties_summary["votes_numer"].iloc[y]
    #Add rows - divide party votes by natural digits
    for i in range(in_mandates_in_district):
        data.append(
            [party_name, votes_number/(i+1)])

#Create Dhondt Data Frame
df_Dhondt = pd.DataFrame(data,columns=["party_name", "votes_number"])
#Sort by Votes number
df_Dhondt.sort_values(by=["votes_number"], ascending=False, inplace=True)
#Limit df to number of mandates in district
df_Dhondt = df_Dhondt.head(in_mandates_in_district)
#Group by Party Name
df_Dhondt = df_Dhondt.groupby(['party_name'])['party_name'].count()


df_winners=pd.DataFrame()
#Loop through all parties
for y in range(df_Dhondt.shape[0]):

    #Get Party name and votes
    party_name = df_Dhondt.axes[0][0]
    mandates_number = df_Dhondt.iloc[0]
    df_party = in_df_data_party_details[in_df_data_party_details['committee_name'] == party_name]
    df_party.sort_values(by=["person_votes_number"], ascending=False, inplace=True)
    df_party = df_party.head(mandates_number)

    df_winners=pd.concat([df_winners,df_party])



#Get list of district
dict_electoral_districts = get_electoral_districts_links(in_webdriver_path = webdriver_path, in_url= url_elecion)

#Loop through all districts
for item in dict_electoral_districts:
    #Get link to current district
    current_link = dict_electoral_districts[item]
    #Get details from current district
    dict_electoral_district_information = get_electoral_districts_details(in_webdriver_path = webdriver_path, in_url= current_link)

    save_df_to_excel("Input\\" + item,dict_electoral_district_information)


#Initialize empty DataFrame
df_winners = pd.DataFrame()

#Loop through all districts
for item in dict_electoral_districts:
    #Get link to current district
    current_link = dict_electoral_districts[item]
    #Get details from current district
    dict_electoral_district_information = get_electoral_districts_details(in_webdriver_path = webdriver_path, in_url= current_link)

    df_data_party_summary = dict_electoral_district_information["df_data_party_summary"]
    df_winner = df_data_party_summary[df_data_party_summary.person_votes_number == df_data_party_summary.person_votes_number.max()]
    if df_winners.empty == True:
        df_winners = df_winner
    else:
        df_winners = pd.concat([df_winners, df_winner])


print("S")



def get_data_from_league(url, webdriver_path):
    #Create webdriver object
    #driver = webdriver.Chrome(webdriver_path)
    service = Service(webdriver_path)
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)


    driver.get(url)
    #print(driver.title)

    #Get table
    table_id = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CLASS_NAME, "ui-table__body")))
    #Get Rows
    rows = table_id.find_elements(By.CLASS_NAME, "ui-table__row  ")

    #Create datatable
    column_names = ["county", "league", "name", "position", "matches_played", "wins", "percentage", "matches_all",
                    "points_left","club_points", "next_club_points", "points_surplus", "points_to_champion"]
    df = pd.DataFrame(columns=column_names)

    #Get first 3 teams
    counter = 0
    for row in rows:
        club_position = row.find_element(By.CLASS_NAME, "tableCellRank")
        club_name = row.find_element(By.CLASS_NAME, "tableCellParticipant__name")
        cell_values = row.find_elements(By.XPATH, './/span[@class = " table__cell table__cell--value   "]')
        club_matches = cell_values[0]
        club_wins = cell_values[1]
        percentage = round(int(club_wins.text)/int(club_matches.text),2)
        matches_all = (len(rows)-1) *2
        points_left = (matches_all- int(club_matches.text)) * 3
        try:
            club_points = row.find_element(By.XPATH,
                                        './/span[@class = " table__cell table__cell--value  table__cell--points "]').text
        except Exception as e:
            club_points = 40
        try:
            #If club is playing game this element is different
            next_club_points = rows[counter+1].find_element(By.XPATH,
                                        './/span[@class = " table__cell table__cell--value  table__cell--points "]').text
        except Exception as e:
            next_club_points = club_points

        points_surplus = int(club_points) - int(next_club_points)
        points_to_champion = points_left -points_surplus

        #Save data in list row
        list_row = [url.split('/')[-2],  #country
                    url.split('/')[-1],  #league
                    club_name.text,  #name
                    club_position.text,  #position
                    club_matches.text,  #matches_played
                    club_wins.text,  #wins
                    percentage,  #percentage
                    matches_all,  #matches_all
                    points_left,  #points_left
                    club_points,  # club_points
                    next_club_points,  # next_club_points
                    points_surplus,
                    points_to_champion
                    ]

        df.loc[len(df)] = list_row

        #If more than 2 exit method
        if counter>=2:
            break
        else:
            counter+=1

    return df


def append_to_excel(fpath, df):
    #If excel exist read sheet as DataTable
    if (os.path.exists(fpath)):
        df_current=pd.read_excel(fpath)
    #If excel does not exist create new DataTable
    else :
        df_current=pd.DataFrame()

    df_concat=pd.concat([df,df_current])
    df_concat.to_excel(fpath,index=False)


def parse_Flashscore():
    # load excel
    excel_input = openpyxl.load_workbook(input_file_name)
    sheet = excel_input.active

    #Delete current output file
    if os.path.exists(output_file_name):
      os.remove(output_file_name)

    # iterate through excel and display data
    for i in range(2, sheet.max_row + 1):
        # Get data from excel
        active_cell_value = bool(sheet.cell(row=i, column=3).value)
        done_cell_value = bool(sheet.cell(row=i, column=4).value)
        date_cell_value = sheet.cell(row=i, column=5).value
        url = sheet.cell(row=i, column=1).value

        # if already done go to next cell
        #Not used right now
        #if done_cell_value:
        #   continue

        # if not active and was deactivated less than 20 days ago go to next cell
        if not active_cell_value and abs(date_cell_value - datetime.datetime.today()).days < 20:
            continue

        try:
            df = get_data_from_league(url, webdriver_path)
        except Exception as e:
            sheet.cell(row=i, column=2).value = "???"
            continue
        # Sort values by percentage
        df.sort_values(by=['percentage'], ascending=False)
        # Save high_score
        high_score = df.at[0, 'percentage']
        append_to_excel(output_file_name, df)

        sheet.cell(row=i, column=2).value = high_score
        # If high_score less than 0.65 - deactivate
        if high_score < 0.7:
            #Set active cell as False
            sheet.cell(row=i, column=3).value = False
            sheet.cell(row=i, column=5).value = datetime.datetime.today()

        # mark as Done
        #sheet.cell(row=i, column=4).value = True

        # Save progress in excel
        excel_input.save(input_file_name)

def transform_output_data():
    # Get output data as DataTable
    df_current = pd.read_excel(output_file_name)

    # Filter only rows with matches_played more than 9 and percentage more than 0.79
    df_current = df_current.loc[df_current['matches_played'] > 9]
    df_current = df_current.loc[df_current['percentage'] > 0.79]

    # Saved filtered data in excel
    df_current.to_excel(output_file_name, index=False)

    # Open excel
    excel_output = openpyxl.load_workbook(output_file_name)
    sheet = excel_output.active

    # Adjust column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    excel_output.save(output_file_name)




#Get data from Flashscore and save them in excel file
#parse_Flashscore()

#Transform data in Excel with specific criteria
#transform_output_data()

